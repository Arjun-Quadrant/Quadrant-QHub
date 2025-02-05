import json
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

class DatasourceInfo:
    def __init__(self):
        self.connections = []

class ConnectionInfo:
    def __init__(self, connection_type, connection_string, tables):
        self.connection_type = connection_type
        self.connection_string = connection_string
        self.tables = tables

class TableInfo:
    def __init__(self, table, columns):
        self.table = table
        self.columns = columns

def extract_data_source_info(file_path: str):
    datasource_info = DatasourceInfo()
    xml_doc = load_xml_document(file_path)
    data_source = get_data_source_element(xml_doc)
    named_connection_info = get_named_connection_elements(data_source)
    for connection in named_connection_info:
        connection_info = get_connection_element(connection)
        connection_type = get_connection_type(connection_info)
        connection_string = build_connection_string(connection_type, connection, connection_info)
        table_column_mapping = extract_table_column_mapping(xml_doc)
        all_tables = build_table_info_list(table_column_mapping)
        datasource_info.connections.append(ConnectionInfo(connection_type, connection_string, all_tables))
    return datasource_info

def load_xml_document(file_path: str):
    return ET.parse(file_path)

def get_data_source_element(xml_doc: ET.ElementTree):
    data_source = xml_doc.find(".//datasource")
    if data_source is None:
        raise ValueError("No datasource found in the Tableau file.")
    return data_source

def get_named_connection_elements(data_source: ET.Element):
    return data_source.findall(".//connection/named-connections/named-connection")

def get_connection_element(connection: ET.Element):
    return connection.find(".//connection")

def get_connection_type(connection_info: ET.Element):
    return connection_info.get("class", "Unknown")

def build_connection_string(connection_type: str, connection: ET.Element, connection_info: ET.Element):
    if connection_type == "sqlserver":
        return (f"sqlserver; Server: {connection.get('caption')}; "
                f"Database: {connection_info.get('dbname')}; "
                f"Authentication: {connection_info.get('authentication')}; "
                f"Require SSL: {'Yes' if connection_info.get('sslmode') == 'require' else 'No'}")
    return connection_info.get("filename", "Unknown")

def extract_table_column_mapping(xml_doc: ET.ElementTree):
    table_column_mapping = {}
    columnContainer = xml_doc.find(".//cols")
    columns = columnContainer.findall("./map")
    for col in columns:
        info = col.get("value")
        if info:
            table_name, _, column_name = re.split(r'(\]\.\[)', info)
            table_name = table_name.strip('[]')
            column_name = column_name.strip('[]')
            table_column_mapping.setdefault(table_name, []).append(column_name)
    return table_column_mapping

def build_table_info_list(table_column_mapping: dict):
    return [TableInfo(table, ", ".join(columns)).__dict__ for table, columns in table_column_mapping.items()]

def save_connection_info_to_json_and_excel(data_source_info: DatasourceInfo, json_file_path: str, excel_file_path: str):
    save_to_json(data_source_info, json_file_path)
    save_to_excel(data_source_info, excel_file_path)

def save_to_json(data_source_info: DatasourceInfo, json_file_path: str):
    with open(json_file_path, 'w') as json_file:
        json.dump([vars(conn) for conn in data_source_info.connections], json_file, indent=4)

def save_to_excel(data_source_info: DatasourceInfo, excel_file_path: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ConnectionInfo"
    headers = ["Data Source(s)", "Connection Info", "Data Table(s)", "Column(s)"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    num_connections = 0
    for connection in data_source_info.connections:
        num_connections = num_connections + 1
        num_cells_to_merge = 0
        for table in connection.tables:
            sheet.append([connection.connection_type, connection.connection_string, table["table"], table["columns"]])
            num_cells_to_merge = num_cells_to_merge + 1
        sheet.merge_cells(start_row=num_connections + 1, start_column=1, end_row=num_connections + num_cells_to_merge, end_column=1)
        sheet.merge_cells(start_row=num_connections + 1, start_column=2, end_row=num_connections + num_cells_to_merge, end_column=2)
        sheet[num_connections + 1][0].alignment = Alignment(horizontal="center", vertical="center")
        sheet[num_connections + 1][1].alignment = Alignment(horizontal="center", vertical="center")
    
    adjust_column_widths(sheet)
    workbook.save(excel_file_path)

def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2