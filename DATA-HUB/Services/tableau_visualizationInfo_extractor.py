import json
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from tableau_api_lib import TableauServerConnection
from collections import namedtuple
from openpyxl import Workbook
import xml.etree.ElementTree as ET
import os
import time

class VisualizationInfo:
    def __init__(self, worksheet_name: str, viz_title: str, viz_type: str, tables_used: str, columns_used: str):
        self.worksheet_name = worksheet_name
        self.viz_title = viz_title
        self.viz_type = viz_type
        self.tables_used = tables_used
        self.column_used = columns_used

def get_workbook_name(file_path: str):
    file_name = os.path.basename(file_path)
    workbook_name, _ = os.path.splitext(file_name)
    return workbook_name

def get_viz_title(worksheet: ET.Element):
    title = worksheet.find(".//run")
    if title is None:
        return "No Title"
    return title.text

def get_viz_type(worksheet: ET.Element):
    vizExists = worksheet.find(".//datasource")
    if vizExists is None:
        return "No Visualization"
    return worksheet.find(".//mark").get("class")

def get_tables_and_columns(workbook_name: str, worksheet_name: str, connection: TableauServerConnection):
    query = f"""
    {{
        workbooks (filter: {{ name: "{workbook_name}"}}) {{
            sheets (filter: {{ name: "{worksheet_name}"}}) {{
                upstreamTables {{
                    name
                }}
                upstreamColumns {{
                    name
                }}
            }}
        }}
    }}
    """
    time.sleep(10)
    response = connection.metadata_graphql_query(query=query)
    response_json = response.json()

    tables_as_json = response_json["data"]["workbooks"][0]["sheets"][0]["upstreamTables"]
    columns_as_json = response_json["data"]["workbooks"][0]["sheets"][0]["upstreamColumns"]
    tablesUsed  = ", ".join(list(dict.values())[0] for dict in tables_as_json)
    columnsUsed = ", ".join(list(dict.values())[0] for dict in columns_as_json)
    return (tablesUsed, columnsUsed)

def save_to_excel(visualization_info_list: list, excel_file_path: str):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "VisualizationInfo"
    add_excel_headers(worksheet)
    populate_excel_data(worksheet, visualization_info_list)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2
    workbook.save(excel_file_path)

def add_excel_headers(worksheet: Worksheet):
    headers = ["Worksheet Name", "Visualization Title", "Visualization Type", "Tables Used", "Columns Used"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

def populate_excel_data(worksheet: Worksheet, visualization_info_list: list):
    for i, vis_info in enumerate(visualization_info_list, start=2):
        worksheet.cell(row=i, column=1, value=vis_info.WorksheetName)
        worksheet.cell(row=i, column=2, value=vis_info.VisualizationTitle)
        worksheet.cell(row=i, column=3, value=vis_info.VisualizationType)
        worksheet.cell(row=i, column=4, value=vis_info.TablesUsed)
        worksheet.cell(row=i, column=5, value=vis_info.ColumnsUsed)

def save_to_json(visualization_info_list: list, json_file_path: str):
    with open(json_file_path, "w", encoding="utf-8") as json_file:
        json.dump([info._asdict() for info in visualization_info_list], json_file, indent=4)

def extract_viz_info(workbook_file_path: str, packaged_workbook_file_path: str):
    config = {
    "tableau_prod": {
        "server": "https://10ay.online.tableau.com",
        "api_version": "3.22",
        "personal_access_token_name": "PLACEHOLDER",
        "personal_access_token_secret": "PLACEHOLDER",
        "site_name": "arjun-at-quadrant",
        "site_url": "arjun-at-quadrant"
        }
    }
    conn = TableauServerConnection(config_json=config, env="tableau_prod")
    conn.sign_in()
    project_name = "Demo"
    project_description = "This project contains workbooks for metadata extraction."
    project_info = conn.create_project(project_name=project_name, project_description=project_description)
    project_info_json = project_info.json()
    project_id = project_info_json["project"]["id"]
    workbook_name = get_workbook_name(packaged_workbook_file_path)
    conn.publish_workbook(workbook_file_path=packaged_workbook_file_path, workbook_name=workbook_name, project_id=project_id, connection_username="PLACEHOLDER", connection_password="PLACEHOLDER")
    workbook_xml_doc = ET.parse(workbook_file_path)
    workbook_worksheets = workbook_xml_doc.findall(".//worksheet")
    viz_info_list = []
    for worksheet in workbook_worksheets:
        worksheet_name = worksheet.get("name")
        viz_title = get_viz_title(worksheet)
        viz_type = get_viz_type(worksheet)
        tables_used = None
        columns_used = None
        if (viz_type == "No Visualization"):
            tables_used = "No Tables Used"
            columns_used = "No Columns Used"
        else:
            (tables_used, columns_used) = get_tables_and_columns(workbook_name=workbook_name, worksheet_name=worksheet_name, connection=conn)
        viz_info = VisualizationInfo(worksheet_name, viz_title, viz_type, tables_used, columns_used)
        viz_info_list.append(viz_info)

def save_viz_info_to_json_and_excel(viz_info: str, json_file_path: str, excel_file_path: str):
    save_to_json(viz_info, json_file_path)
    save_to_excel(viz_info, excel_file_path)